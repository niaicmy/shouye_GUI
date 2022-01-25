import os
import shutil
import datetime
import pandas as pd
import openpyxl
import pymssql
import json

# 行号是5  开始 依次5，6，7，。。。。。
# 入院途径有两个 手动删一个没有数据的,  当前病区, 入院情况 、冷沉淀（U）没有，  无手术一类，应该是单独的库
# 入院转入医疗机构名称  医嘱转院下第二项 转入社区、乡镇卫生院 不知道是不是这个
# 需要拆分的是时间, 疾病列表,

predefined_columns = ['姓名', '住院编号', '付款方式', '健康卡号', '入院当次', '病案编号',
                      '性别', '出生日期', '年龄_岁', '国籍',
                      '年龄_月', '年龄_天', '新生儿出生体重', '新生儿入院体重', '出生地址县', '户口地址县',
                      '民族', '身份证号', '职业', '婚姻状态', '现住址县', '现住址镇',
                      '现住址电话', '现住址邮编', '户口地址县', '户口地址镇', '户口邮编', '工作单位县', '工作单位镇',
                      '电话', '工作单位邮编', '联系人姓名', '联系人关系', '联系人地址县', '联系人地址镇',
                      '联系人电话', '入院途径', '入院时间',
                      '入院病区', '入院病室', '转科科别', '出院时间',
                      '当前病区', '当前病室', '实际天数', '门诊诊断', '门诊诊断编码',
                      '入院诊断', '入院诊断编码', '出现危重',
                      '诊断名称列表', '诊断编码列表',
                      '损伤中毒外部因素', '损伤中毒外部因素编码',
                      '病理诊断', '病理诊断编码', '病理号',
                      '有无过敏', '药物过敏', '尸检否',
                      '血型', 'RH',
                      '输血红细胞', '输血血小板', '输血血浆', '输血全血', '输血自体血回输', '输血白蛋白',
                      '输血其它', '输血反应',
                      '是否随诊', '随诊期限周', '随诊期限月', '随诊期限年',
                      '科主任', '主任医生', '主诊医生', '主治医生', '住院医师',
                      '责任护士', '进修医生', '实习医生', '编码员',
                      '病案质量', '质控医师', '质控护士', '质控日期',
                      '临床路径管理', '完成临床路径', '退出临床原因', '是否变异', '变异原因',
                      '出院方式', '转入医疗机构名称', '入院转入医疗机构名称',
                      'CT', 'PETCT', '双源CT', 'B超',
                      'X片', '超声心动图', 'MRI', '同位素检查',
                      '再住院计划', '再住院目的',
                      '颅脑损伤昏迷入院前_天', '颅脑损伤昏迷入院前_小时', '颅脑损伤昏迷入院前_分钟', '颅脑损伤昏迷入院后_天', '颅脑损伤昏迷入院后_小时', '颅脑损伤昏迷入院后_分钟',
                      '费用总额', '自付金额'
                      ]


# 包装成函数:
def load_json_file(json_path):
    try:
        json_file = open(json_path, 'r', encoding='gbk')
        file_dict = json.load(json_file)
    except UnicodeError:
        json_file = open(json_path, 'r', encoding='utf-8')
        file_dict = json.load(json_file)

    return file_dict


def target_mkdir():
    now_time = datetime.datetime.now().strftime("%Y-%m-%d")
    # print(now_time)
    file_path = None
    # 获取到当前文件的目录，并检查是否有 directory_time 文件夹，如果不存在则自动新建 directory_time 文件
    try:
        file_path = os.getcwd() + "\\" + now_time
        print(file_path)
        # 判断是否已经存在该目录
        if not os.path.exists(file_path):
            # 目录不存在，进行创建操作
            os.makedirs(file_path)  # 使用os.makedirs()方法创建多层目录
            print("目录新建成功：" + file_path)
            return file_path
        else:
            print("目录已存在！！！")
            return file_path
    except BaseException as msg:
        print("新建目录失败：" + str(msg))

    return file_path


# 读取员工表目录
def staff_records(records):
    # 读取 excel 数据
    df_source = pd.read_excel(records)

    # 从源数据中读取需要的数据
    data_no = df_source.loc[:, '编号'].values  # 读所有行的title以及data列的值，这里需要嵌套列表
    data_name = df_source.loc[:, '名称'].values  # 读所有行的title以及data列的值，这里需要嵌套列表
    staff_records_dict = dict(zip(data_no, data_name))
    # print(staff_records_dict)

    return staff_records_dict


# pandas 教程
# https://www.cnblogs.com/nxf-rabbit75/p/10105271.html
# https://www.cnblogs.com/helenlee01/p/12617481.html
def make_excel(source_path, staff_records_path, sql_config_path):
    staff_dict = staff_records(staff_records_path)
    target_data_path = target_mkdir()

    # 获取数据库 emr 库连接参数
    sql_config = load_json_file(sql_config_path)
    emr_config = sql_config['emr_config']
    # 打开数据库 获取游标
    db = pymssql.connect(**emr_config)
    cursor = db.cursor()

    # 读取 excel 数据
    df_source = pd.read_excel(source_path)
    # 统一数据处理-替换数据中的内容
    # df_source.fillna(r'')
    df_source.replace(r'新津县', r'新津区', inplace=True)
    df_source.replace(r'本人', r'本人或户主', inplace=True)
    df_source.replace(r'住院部', r'其他', inplace=True)

    # todo: 循环读取每一行数据 做出模板，还要做一个检测人员病案有没有的功能
    # 循环读取从源数据中读取需要的数据
    # 获取行数
    # print(df_source.shape[0])
    # print(len(df_source))
    for index in range(df_source.shape[0]):
        # todo: 单独设定 index 为了读取指定人员病案
        # index = 23
        personal_data = df_source.loc[index, predefined_columns].values  # 读所有行的title以及data列的值，这里需要嵌套列表
        # print(len(personal_data))
        # 执行费用数据读取sql，提取费用 数据库语句在 config 中配置
        sql_select = sql_config['sql_select_config']['emr_data_select'].format(personal_data[1])
        cursor.execute(sql_select)
        results = cursor.fetchall()  # fetchall()是接收全部的返回结果行
        # print(results)
        # 生成费用字典
        dict_key = []
        dict_value = []
        for data in results:
            dict_key.append(data[0])
            dict_value.append(data[1])

        expense_dict = dict(zip(dict_key, dict_value))

        write_data(target_data_path, personal_data, staff_dict, expense_dict)

    # 数据写入完后关闭数据库连接
    cursor.close()
    db.close()


def write_data(target_path, personal_data, staff_dict, expense_dict):
    target_file_name = target_path + '\\' + str(personal_data[0]) + str(personal_data[1]) + r'.xlsx'
    # 复制报表模板文件到目标文件目录，并重命名
    shutil.copyfile(r'template_file.xlsx', target_file_name)
    w_book = openpyxl.load_workbook(target_file_name)
    # print(book.worksheets)
    sheet = w_book.active
    # sheet = w_book['XXB']
    # sheet_names = work_book.get_sheet_names()
    # print(sheet_names)
    # sheet = work_book.get_sheet_by_name(sheet_names[0])

    # 访问A列4行的单元格，不存在则创建
    # c = sheet['A4']
    # 还有Worksheet.cell()方法，赋值（4，1）值为10
    # d = sheet.cell(row=4, column=1, value=10)
    # sheet.cell(row=2, column=2).value = 2
    # p = (str(data[2]), str(data[3]), str(data[4]), str(data[0]))
    # 以下为数据写入
    # 第六行
    sheet['D6'].value = str(personal_data[2])
    sheet['F6'].value = int(personal_data[3])
    sheet['J6'].value = int(personal_data[4])
    sheet['L6'].value = int(personal_data[5])
    # 第七行
    sheet['D7'].value = str(personal_data[0])
    sheet['F7'].value = str(personal_data[6])
    sheet['J7'].value = personal_data[7].strftime("%Y%m%d")
    sheet['L7'].value = str(personal_data[8])
    sheet['N7'].value = str(personal_data[9])
    # 第八行
    # sheet['D8'].value = str(data[10])
    if personal_data[10] == 'nan' or personal_data[10] == 0:
        sheet['D8'].value = ''
    else:
        sheet['D8'].value = str(personal_data[10])

    # sheet['F8'].value = str(data[11])
    if personal_data[11] == 'nan' or personal_data[11] == 0:
        sheet['F8'].value = ''
    else:
        sheet['F8'].value = str(personal_data[11])

    # sheet['H8'].value = str(data[12])
    # print(data[12])
    if str(personal_data[12]) == 'nan' or personal_data[12] == 0:
        sheet['H8'].value = ''
    else:
        sheet['H8'].value = str(personal_data[12])

    # sheet['J8'].value = str(data[13])
    if personal_data[13] == 'nan' or personal_data[13] == 0.0:
        sheet['J8'].value = ''
    else:
        sheet['J8'].value = str(personal_data[13])

    sheet['L8'].value = str(personal_data[14])
    sheet['N8'].value = str(personal_data[15])
    # 第九行
    sheet['D9'].value = str(personal_data[16])
    sheet['F9'].value = str(personal_data[17])
    # sheet['J9'].value = str(personal_data[18])
    if personal_data[18] == '退(离)休人员':
        sheet['J9'].value = '退（离）休人员'
    else:
        sheet['J9'].value = str(personal_data[18])

    sheet['L9'].value = str(personal_data[19])
    sheet['N9'].value = str(personal_data[20])
    sheet['O9'].value = str(personal_data[21])
    # 第十行
    sheet['D10'].value = str(personal_data[29])
    sheet['F10'].value = int(personal_data[23])
    sheet['I10'].value = str(personal_data[24])
    sheet['J10'].value = str(personal_data[25])
    sheet['L10'].value = int(personal_data[26])
    sheet['N10'].value = str(personal_data[27])
    sheet['O10'].value = str(personal_data[28])
    # 第十一行
    sheet['D11'].value = str(personal_data[29])
    sheet['F11'].value = int(personal_data[30])
    sheet['J11'].value = str(personal_data[31])
    # sheet['L11'].value = str(data[32])
    if str(personal_data[32]) == '本人' or personal_data[32] == 0:
        sheet['L11'].value = '本人或户主'
    else:
        sheet['L11'].value = str(personal_data[32])

    sheet['N11'].value = str(personal_data[33])
    sheet['O11'].value = str(personal_data[34])
    # 第十二行
    sheet['D12'].value = str(personal_data[35])
    if str(personal_data[35]) == 'nan' or personal_data[35] == 0:
        sheet['D12'].value = str(personal_data[29])
    else:
        sheet['D12'].value = int(personal_data[35])

    sheet['F12'].value = str(personal_data[36])
    sheet['J12'].value = personal_data[37].strftime("%Y%m%d")
    sheet['L12'].value = personal_data[37].strftime("%H")
    sheet['N12'].value = personal_data[37].strftime("%M")
    # 第十三行
    # sheet['D13'].value = str(data[38])
    if str(personal_data[38]) == '住院部' or personal_data[38] == 0:
        sheet['D13'].value = '其他'
    else:
        sheet['D13'].value = str(personal_data[38])

    sheet['F13'].value = str(personal_data[39])
    # sheet['H13'].value = str(data[40])
    # print(data[40])
    if str(personal_data[40]) == 'nan' or personal_data[40] == 0:
        sheet['H13'].value = '-'
    else:
        sheet['H13'].value = str(personal_data[40])

    sheet['J13'].value = personal_data[41].strftime("%Y%m%d")
    sheet['L13'].value = personal_data[41].strftime("%H")
    sheet['N13'].value = personal_data[41].strftime("%M")
    # 第十四行
    # sheet['D14'].value = str(data[42])
    if str(personal_data[42]) == '住院部' or personal_data[42] == 0:
        sheet['D14'].value = '其他'
    else:
        sheet['D14'].value = str(personal_data[42])

    sheet['F14'].value = str(personal_data[43])
    sheet['H14'].value = str(personal_data[44])
    sheet['J14'].value = str(personal_data[45]).split(',')[0]
    sheet['L14'].value = str(personal_data[46]).split(',')[0]
    # 第十五行
    sheet['D15'].value = str(personal_data[47]).split(',')[0]
    sheet['H15'].value = str(personal_data[48]).split(',')[0]
    sheet['N15'].value = str(personal_data[49])
    # 第十八行
    # 50,51: '诊断名称列表', '诊断编码列表', str(data[50]).split(',')[0]
    disease_name = str(personal_data[50]).split(',')
    disease_no = str(personal_data[51]).split(',')
    t_len = len(disease_name)
    for index in range(t_len):
        # print(index)
        if index <= 7:
            sheet['D{0}'.format(18 + index)].value = disease_name[index]
            sheet['F{0}'.format(18 + index)].value = disease_no[index]
            sheet['H{0}'.format(18 + index)].value = '有'
        elif 8 <= index < 16:
            sheet['J{0}'.format(18 + index - 8)].value = disease_name[index]
            sheet['K{0}'.format(18 + index - 8)].value = disease_no[index]
            sheet['N{0}'.format(18 + index - 8)].value = '有'

    if t_len <= 7:
        sheet['D{0}'.format(18 + t_len)].value = '-'
        sheet['F{0}'.format(18 + t_len)].value = '-'
        sheet['H{0}'.format(18 + t_len)].value = '-'
    elif t_len <= 16:
        sheet['J{0}'.format(18 + t_len - 8)].value = '-'
        sheet['K{0}'.format(18 + t_len - 8)].value = '-'
        sheet['N{0}'.format(18 + t_len - 8)].value = '-'

    # 第二十六行
    # sheet['D26'].value = str(data[52])
    if str(personal_data[52]) == 'nan' or personal_data[52] == 0:
        sheet['D26'].value = '-'
    else:
        sheet['D26'].value = str(personal_data[52])

    # sheet['K26'].value = str(data[53])
    if str(personal_data[53]) == 'nan' or personal_data[53] == 0:
        sheet['K26'].value = '-'
    else:
        sheet['K26'].value = str(personal_data[53])
    # 第二十七行
    # sheet['D27'].value = str(data[54])
    if str(personal_data[54]) == 'nan' or personal_data[54] == 0:
        sheet['D27'].value = '-'
    else:
        sheet['D27'].value = str(personal_data[54])

    # sheet['J27'].value = str(data[55])
    if str(personal_data[55]) == 'nan' or personal_data[55] == 0:
        sheet['J27'].value = '-'
    else:
        sheet['J27'].value = str(personal_data[55])

    # sheet['M27'].value = str(data[56])
    if str(personal_data[56]) == 'nan' or personal_data[56] == 0:
        sheet['M27'].value = '-'
    else:
        sheet['M27'].value = str(personal_data[56])

    # 第二十八行
    sheet['D28'].value = str(personal_data[57])
    # sheet['J28'].value = str(data[58])
    if str(personal_data[58]) == 'nan' or personal_data[58] == 0:
        sheet['J28'].value = '-'
    else:
        sheet['J28'].value = str(personal_data[58])

    # sheet['M28'].value = str(data[59])
    if str(personal_data[59]) == 'nan' or personal_data[59] == 0:
        sheet['M28'].value = '-'
    else:
        sheet['M28'].value = str(personal_data[59])
    # 第二十九行
    sheet['D29'].value = str(personal_data[60])
    sheet['K29'].value = str(personal_data[61])
    # 第三十行
    sheet['D30'].value = int(personal_data[62])
    sheet['F30'].value = int(personal_data[63])
    sheet['H30'].value = int(personal_data[64])
    sheet['J30'].value = int(personal_data[65])
    sheet['M30'].value = int(personal_data[66])
    sheet['O30'].value = int(personal_data[67])
    # 第三十一行
    sheet['D31'].value = 0
    # sheet['F31'].value = str(data[68])
    if str(personal_data[68]) == 'nan' or personal_data[68] == 0:
        sheet['F31'].value = 0
    else:
        sheet['F31'].value = str(personal_data[68])
    # sheet['J31'].value = str(personal_data[69])
    if str(personal_data[69]) == 'nan' or personal_data[69] == 0:
        sheet['J31'].value = '未输'
    else:
        sheet['J31'].value = str(personal_data[69])

    # 第三十二行
    sheet['D32'].value = str(personal_data[70])

    # sheet['J32'].value = str(data[71])
    if str(personal_data[71]) == 'nan' or personal_data[71] == 0:
        sheet['J32'].value = ''
    else:
        sheet['J32'].value = str(personal_data[71])

    # sheet['L32'].value = str(data[72])
    if str(personal_data[72]) == 'nan' or personal_data[72] == 0:
        sheet['L32'].value = ''
    else:
        sheet['L32'].value = str(personal_data[72])

    # sheet['N32'].value = str(data[73])
    if str(personal_data[73]) == 'nan' or personal_data[73] == 0:
        sheet['N32'].value = ''
    else:
        sheet['N32'].value = str(personal_data[73])
    # 第三十三行
    # print(data[74])
    # print(type(data[74]))
    sheet['D33'].value = staff_dict.get(int(personal_data[74]), '-')
    sheet['G33'].value = staff_dict.get(int(personal_data[75]), '-')
    sheet['I33'].value = staff_dict.get(int(personal_data[76]), '-')
    sheet['K33'].value = staff_dict.get(int(personal_data[77]), '-')
    sheet['N33'].value = staff_dict.get(int(personal_data[78]), '-')

    # 第三十四行 -- 2022年新增

    # 第三十五行
    sheet['D35'].value = staff_dict.get(int(personal_data[79]), '-')
    sheet['G35'].value = staff_dict.get(int(personal_data[80]), '-')
    sheet['K35'].value = staff_dict.get(int(personal_data[81]), '-')
    # sheet['N34'].value = str(data[82])
    if str(personal_data[82]) == 'nan' or personal_data[82] == 0:
        sheet['N35'].value = ''
    else:
        sheet['N35'].value = str(personal_data[82])

    # 第三十六行 -- 2022年新增

    # 第三十七行
    sheet['D37'].value = str(personal_data[83])
    sheet['G37'].value = str(personal_data[84])
    sheet['K37'].value = str(personal_data[85])
    sheet['N37'].value = personal_data[86].strftime("%Y%m%d")
    # 第四十一行  手术填报
    sheet['C41'].value = '-'
    sheet['D41'].value = '-'
    sheet['E41'].value = '-'
    sheet['F41'].value = '-'
    sheet['G41'].value = '-'
    sheet['I41'].value = '-'
    sheet['J41'].value = '-'
    sheet['K41'].value = '-'
    sheet['L41'].value = '-'
    sheet['M41'].value = '-'
    sheet['N41'].value = '-'
    sheet['O41'].value = '-'
    # 第四十八行
    sheet['E48'].value = str(personal_data[87])
    # sheet['G46'].value = str(data[88])
    if str(personal_data[88]) == 'nan' or personal_data[88] == 0:
        sheet['G48'].value = '-'
    else:
        sheet['G48'].value = str(personal_data[88])
    # sheet['J46'].value = str(data[89])
    if str(personal_data[89]) == 'nan' or personal_data[89] == 0:
        sheet['J48'].value = '-'
    else:
        sheet['J48'].value = str(personal_data[89])
    # sheet['L46'].value = str(data[90])
    if str(personal_data[90]) == 'nan' or personal_data[90] == 0:
        sheet['L48'].value = '-'
    else:
        sheet['L48'].value = str(personal_data[90])
    # sheet['N46'].value = str(data[91])
    if str(personal_data[91]) == 'nan' or personal_data[91] == 0:
        sheet['N48'].value = '-'
    else:
        sheet['N48'].value = str(personal_data[91])
    # 第四十九行
    sheet['E49'].value = str(personal_data[92])
    if str(personal_data[92]) == '医嘱转院':
        sheet['M49'].value = str(personal_data[93])
    if str(personal_data[92]) == '医嘱转社区/乡镇':
        sheet['M50'].value = str(personal_data[94])
    # 第五十一行
    # sheet['E49'].value = str(data[95])
    if str(personal_data[95]) == 'nan' or personal_data[95] == 0:
        sheet['E51'].value = '未做'
    else:
        sheet['E51'].value = str(personal_data[95])
    # sheet['H49'].value = str(data[96])
    if str(personal_data[96]) == 'nan' or personal_data[96] == 0:
        sheet['H51'].value = '未做'
    else:
        sheet['H51'].value = str(personal_data[96])
    # sheet['K49'].value = str(data[97])
    if str(personal_data[97]) == 'nan' or personal_data[97] == 0:
        sheet['K51'].value = '未做'
    else:
        sheet['K51'].value = str(personal_data[97])
    # sheet['N49'].value = str(data[98])
    if str(personal_data[98]) == 'nan' or personal_data[98] == 0:
        sheet['N51'].value = ''
    else:
        sheet['N51'].value = str(personal_data[98])
    # 第五十二行
    # sheet['E50'].value = str(data[99])
    if str(personal_data[99]) == 'nan' or personal_data[99] == 0:
        sheet['E52'].value = ''
    else:
        sheet['E52'].value = str(personal_data[99])
    # sheet['H50'].value = str(data[100])
    if str(personal_data[100]) == 'nan' or personal_data[100] == 0:
        sheet['H52'].value = '未做'
    else:
        sheet['H52'].value = str(personal_data[100])
    # sheet['K50'].value = str(data[101])
    if str(personal_data[101]) == 'nan' or personal_data[101] == 0:
        sheet['K52'].value = '未做'
    else:
        sheet['K52'].value = str(personal_data[101])
    # sheet['N50'].value = str(data[102])
    if str(personal_data[102]) == 'nan' or personal_data[102] == 0:
        sheet['N52'].value = '未做'
    else:
        sheet['N52'].value = str(personal_data[102])
    # 第五十三行
    sheet['E53'].value = str(personal_data[103])
    # sheet['M51'].value = str(data[104])
    if str(personal_data[104]) == 'nan' or personal_data[104] == 0:
        sheet['M53'].value = '-'
    else:
        sheet['M53'].value = str(personal_data[104])
    # 第五十四行
    sheet['E54'].value = int(personal_data[105])
    sheet['I54'].value = int(personal_data[106])
    sheet['K54'].value = int(personal_data[107])
    # 第五十五行
    sheet['E55'].value = int(personal_data[108])
    sheet['I55'].value = int(personal_data[109])
    sheet['K55'].value = int(personal_data[110])
    # 以下为费用信息填报
    # # 执行费用数据读取sql，提取费用 数据库语句在 config 中配置
    # sql_select = sql_config_p['sql_select_config']['emr_data_select'].format(personal_data[1])
    # cursor.execute(sql_select)
    # results = cursor.fetchall()  # fetchall()是接收全部的返回结果行
    # # print(results)
    #
    # # 生成费用字典
    # dict_key = []
    # dict_value = []
    # for data in results:
    #     # print(data[0])
    #     # print(data[1])
    #     dict_key.append(data[0])
    #     dict_value.append(data[1])
    # expense_dict = dict(zip(dict_key, dict_value))
    # # expense_dict = {}

    # 第五十七行
    sheet['E57'].value = str(personal_data[111])
    sheet['L57'].value = str(personal_data[112])
    sheet['N57'].value = 0
    # 第五十八行
    sheet['E58'].value = expense_dict.get('综合类_医疗服务费', 0.00)
    sheet['I58'].value = expense_dict.get('综合类_治疗操作费', 0.00)
    sheet['K58'].value = expense_dict.get('综合类_护理费', 0.00)
    sheet['N58'].value = expense_dict.get('综合类_其他费', 0.00)
    # 第五十九行
    sheet['E59'].value = expense_dict.get('诊断类_病理费', 0.00)
    sheet['I59'].value = expense_dict.get('诊断类_实验费', 0.00)
    # 是否 根据 影像费用 来设置 B超，X片 项目
    sheet['K59'].value = expense_dict.get('诊断类_影像费', 0.00)
    sheet['N59'].value = expense_dict.get('诊断类_临床诊断费', 0.00)
    # 第六十行 手术项目
    sheet['E60'].value = 0.00
    sheet['I60'].value = expense_dict.get('治疗类_非临床物理治疗费', 0.00)
    sheet['K60'].value = expense_dict.get('治疗类_临床物理治疗费', 0.00)
    sheet['M60'].value = expense_dict.get('治疗类_麻醉费', 0.00)
    sheet['O60'].value = expense_dict.get('治疗类_手术费', 0.00)
    # 第六十一行
    sheet['E61'].value = expense_dict.get('康复类_康复费', 0.00)
    sheet['I61'].value = expense_dict.get('中医类_中医治疗费', 0.00)
    sheet['K61'].value = expense_dict.get('西药类_西药费', 0.00)
    sheet['M61'].value = expense_dict.get('西药类_抗菌药物费', 0.00)
    # 第六十二行
    sheet['E62'].value = expense_dict.get('中药类_成药费', 0.00)
    sheet['I62'].value = expense_dict.get('中药类_草药费', 0.00)
    sheet['K62'].value = expense_dict.get('血液类_血费', 0.00)
    sheet['N62'].value = expense_dict.get('血液类_白蛋白制品费', 0.00)
    # 第六十三行
    sheet['E63'].value = expense_dict.get('血液类_球蛋白制品费', 0.00)
    sheet['I63'].value = expense_dict.get('血液类_凝血因子类制品费', 0.00)
    sheet['K63'].value = expense_dict.get('血液类_细胞因子类制品费', 0.00)
    sheet['O63'].value = expense_dict.get('耗材类_检查材料费', 0.00)
    # 第六十四行
    sheet['E64'].value = expense_dict.get('耗材类_治疗材料费', 0.00)
    sheet['I64'].value = expense_dict.get('耗材类_手术材料费', 0.00)
    sheet['M64'].value = expense_dict.get('其他类_其他费', 0.00)
    # 第六十五行 人员信息
    # todo: 最后一行人员信息修改位置
    sheet['D65'].value = '罗志平'
    # sheet['F63'].value = '侯锡文'
    sheet['F65'].value = '杨大林'
    sheet['I65'].value = ''
    sheet['K65'].value = '02882517666'
    sheet['M65'].value = '02882517666'
    sheet['O65'].value = datetime.datetime.now().strftime("%Y%m%d")

    # 每次写完数据都要保存文件
    w_book.save(target_file_name)


if __name__ == '__main__':
    # 读取人事档案
    staff_records_name = r'staff_records.xlsx'
    staff_records_path = os.getcwd() + "\\" + staff_records_name

    # 读取出院数据 生成病案文件
    # source_file_name = r'source_data.xlsx'
    source_file_name = r'202106病案数据.xlsx'
    source_data_path = os.getcwd() + "\\" + source_file_name

    # 数据库文件 路径
    sql_config_name = r'database_config.json'
    sql_config_path = os.getcwd() + "\\" + sql_config_name

    make_excel(source_data_path, staff_records_path, sql_config_path)
